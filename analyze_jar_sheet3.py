import openpyxl

xlsx_path = 'experience/PT JAR Andalan Rasa/Soal Test Offline - DA_PT JAR Andalan Rasa_Anggit Djoko.xlsx'

try:
    wb = openpyxl.load_workbook(xlsx_path)
    print(f"Sheet names: {wb.sheetnames}")
    
    # Get sheet 3 (index 2)
    if len(wb.sheetnames) >= 3:
        ws = wb.worksheets[2]
        print(f"\n=== Sheet 3: {ws.title} ===")
        print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
        
        # Print first 5 rows
        rows = list(ws.iter_rows(values_only=True))
        print("First 5 rows:")
        for i, row in enumerate(rows[:5]):
            print(f"Row {i+1}: {row}")
    else:
        print("Less than 3 sheets found")
except Exception as e:
    print(f"Error: {e}")
